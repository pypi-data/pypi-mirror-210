import datetime
import json
from openpyxl import load_workbook

def out_duty():
    workbook_object = load_workbook(filename='总值班2023.xlsx')
    now = datetime.datetime.now()
    shett_object = workbook_object["1226修改"]
    result = shett_object.iter_rows(
        min_row=2, max_row=400, min_col=1, max_col=10, values_only=True)
    employees = [i for i in result if i[0] != None]
    sort_date = sorted(employees, reverse=False, key=lambda x: x[0])
    today = datetime.datetime.now().date()
    today_index = ''
    for i, item in enumerate(sort_date):
        if item[0].date() == today:
            today_index = i
    head_sort_data = sort_date[today_index:today_index+60]


    # 转换为首页的格式
    final_data = []
    for i1 in head_sort_data:
        duty = [i for i in i1 if i != None]
        personnel_on_duty = duty[2:]
        result = {"date": duty[0].strftime(
            "%Y-%#m-%#d"), "lead": duty[1], "person": personnel_on_duty}
        final_data.append(result)


    # 转json
    with open("dutys.js", "w", encoding="utf-8") as f:
        json.dump(final_data, f, ensure_ascii=False)

    with open("dutys.js", "r+", encoding="utf-8") as f:
        content = f.read()
        f.seek(0, 0)
        f.write('var json='+content)

    # 转换为详细内容的格式
    cal_data = []
    for i1 in sort_date:
        duty = [i for i in i1 if i != None]
        personnel_on_cal = "<br>".join(duty[2:])
        lead = duty[1]
        cal_title = f'<span>带班领导：{lead}</span><br>{personnel_on_cal}'
        result = {"start": duty[0].strftime(
            "%Y-%m-%d"), "title": cal_title}
        cal_data.append(result)

    with open("fullcal.js", "w", encoding="utf-8") as f:
        json.dump(cal_data, f, ensure_ascii=False)
