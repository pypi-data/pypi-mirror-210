import os
import json
from openpyxl import Workbook, load_workbook

class ExcelTool:

    @staticmethod
    def json_to_excel(json_data, file_path, sheet_name='Sheet1', overwrite=False):
        """
        json数据写入到excel中
        :param json_data:  json数据
        :param file_path:  文件路径
        :param sheet_name:  sheet名称
        :param overwrite:  是否覆盖，默认不覆盖
        :return:
        """
        if os.path.exists(file_path) and not overwrite:
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        if sheet_name in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists in the file.")
        sheet = workbook.create_sheet(title=sheet_name)

        if isinstance(json_data, str):
            json_data = json.loads(json_data)

        if isinstance(json_data, list):
            for i, item in enumerate(json_data):
                if isinstance(item, dict):
                    row_data = list(item.values())
                    sheet.append(row_data)
                else:
                    sheet.append([item])
        elif isinstance(json_data, dict):
            header = list(json_data.keys())
            sheet.append(header)
            row_data = list(json_data.values())
            sheet.append(row_data)

        workbook.save(file_path)

    @staticmethod
    def excel_to_json(file_path, sheet_name='Sheet1'):
        """
        读取excel数据转回json
        :param file_path:
        :param sheet_name:
        :return:
        """
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        header = data[0]
        json_data = []

        for row in data[1:]:
            json_row = {}
            for i, value in enumerate(row):
                json_row[header[i]] = value
            json_data.append(json_row)

        return json.dumps(json_data)

    @staticmethod
    def read_excel(file_path, sheet_name='Sheet1'):
        """
        读取excel 转化为二维数组
        :param file_path:
        :param sheet_name:
        :return:
        """
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    @staticmethod
    def write_excel(data, file_path, sheet_name='Sheet1',overwrite=False):
        """
        二维数组写入到excel
        :param data: 二维数组数据
        :param file_path: 文件路径
        :param sheet_name: sheet名称
        :param overwrite: 是否覆盖原来文件
        :return:
        """
        if os.path.exists(file_path) and not overwrite:
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()
        sheet = workbook.create_sheet(title=sheet_name)
        for row in data:
            sheet.append(row)
        workbook.save(file_path)

    @staticmethod
    def copy_sheet(source_file, source_sheet, destination_file, destination_sheet):
        """
        复制sheet 支持同一文件和跨文件
        :param source_file:
        :param source_sheet:
        :param destination_file:
        :param destination_sheet:
        :return:
        """
        source_workbook = load_workbook(source_file)
        source_worksheet = source_workbook[source_sheet]

        if source_file == destination_file:
            ws_copy = source_workbook.copy_worksheet(source_worksheet)
            ws_copy.title = destination_sheet
            source_workbook.save(source_file)
            return
        if not os.path.exists(destination_file):
            Workbook().save(destination_file)
        destination_workbook = load_workbook(destination_file)
        if destination_sheet in destination_workbook.sheetnames:
            destination_workbook.remove(destination_workbook[destination_sheet])
        destination_worksheet = source_workbook.copy_worksheet(source_worksheet)
        destination_worksheet.title = destination_sheet
        destination_workbook.save(destination_file)


    @staticmethod
    def delete_sheet(file_path, sheet_name):
        """
        删除指定excel的shell
        :param file_path:
        :param sheet_name:
        :return:
        """
        workbook = load_workbook(file_path)
        sheet_names = workbook.sheetnames
        if sheet_name in sheet_names:
            workbook.remove(workbook[sheet_name])
            workbook.save(file_path)