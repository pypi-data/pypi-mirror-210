import os
import csv
import json

from openpyxl import load_workbook

RELEASE_DATE = "2023-05-28"
COLOR_STYLES = {
    "normal":{
        "name":"标准",
        "colors":{
            "chessboard":"khaki",
            "red-label":"pink",
            "blue-label":"skyblue",
            "chess-bg":"lightyellow",
            "selected-chess-bg":"yellow",
            "blank-feasible-bg":"lightgreen",
            "occupied-feasible-bg":"pink",
            "red-chess-fg":"lightcoral",
            "red-tran-chess-fg":"red",
            "blue-chess-fg":"cornflowerblue",
            "blue-tran-chess-fg":"blue",
            "neutral-chess-fg":"green"
        }
    },
    "dark":{
        "name":"暗色",
        "colors":{
            "chessboard":"grey",
            "red-label":"pink",
            "blue-label":"skyblue",
            "chess-bg":"darkslategrey",
            "selected-chess-bg":"darkgrey",
            "blank-feasible-bg":"green",
            "occupied-feasible-bg":"firebrick",
            "red-chess-fg":"pink",
            "red-tran-chess-fg":"lightcoral",
            "blue-chess-fg":"deepskyblue",
            "blue-tran-chess-fg":"dodgerblue",
            "neutral-chess-fg":"lightgreen"
        }
    }
}
COMPONENT_STYLE = {
    "btn-style":{
        "relief":"groove",
        "bg":"snow"
    },
    "label-style":{
        "relief":"groove"
    }
}

SETTING_PATH = "setting.json"

ARR = tuple[int,int] # 棋子位置数组

def load(path:str):
    with open(path,encoding = "utf-8") as rfile:
        return json.load(rfile)

def write(path:str,data):
    with open(path,"w",encoding = "utf-8") as wfile:
            json.dump(data,wfile)

class MapViewer:
    # 解析位置条件
    @staticmethod
    def parse_location(data:str):
        """
        X[1]&Y[2]&P[2,3]&T[7]
        ->
        [('X',1),('Y',2),('P',2,3),('T',7)]
        """
        data = data.split("&")
        loc = list[tuple]()
        for i in data:
            if not i:
                continue
            command = i.split("[")[0]
            args = [int(j) for j in i[len(command) + 1:-1].split("|")]
            loc.append(tuple([command] + args))
        return loc

    # 解析移动规则
    @staticmethod
    def parse_move(data:str):
        """
        1(Y[1|2]),2(Y[3|4]),3
        ->
        [(1,[('Y',1,2)]),(2,[('Y',3,4)]),(3,)]
        """
        data = data.split(",")
        move = list[tuple]()
        for i in data:
            if not i:
                continue
            if "(" in i and i.endswith(")"):
                move.append((
                    int(i.split("(")[0]),
                    MapViewer.parse_location(i.split("(")[1][:-1])
                ))
            else:
                move.append((int(i),))
        return move

    @staticmethod
    def view(filename:str):
        wb = load_workbook(filename)
        # 处理棋子
        chesses_sheet = wb.get_sheet_by_name("chesses")
        chesses = []
        title = chesses_sheet[1]
        for i in chesses_sheet[2:chesses_sheet.max_row]:
            name = str(i[1].value).strip('"')
            belong = int(i[2].value)
            is_captain = i[3].value == "c"
            move = [MapViewer.parse_move(j) for j in str(i[4].value).split(";")]
            tran_con = MapViewer.parse_location(i[5].value) if i[5].value else []
            tran_move = [MapViewer.parse_move(j) for j in str(i[6].value).split(";")]
            attr = dict([(title[j + 7].value,k.value) for j,k in enumerate(i[7:]) if k.value])
            if len(move) < 2 or len(tran_move) < 2:
                raise TypeError
            chesses.append([name,belong,is_captain,move,tran_con,tran_move,attr])
        # 处理地图
        map_sheet = wb.get_sheet_by_name("map")
        map = []
        rl = int(map_sheet[1][1].value)
        cl = int(map_sheet[1][3].value)
        for i in range(rl):
            map.append([])
            for j in range(cl):
                val = map_sheet[i + 2][j]
                map[-1].append(int(val.value) if val.value else None)
        # 处理特殊规则
        rules_sheet = wb.get_sheet_by_name("rules")
        rules = {}
        rules["tran"] = rules_sheet[2][2].value == "c" # 启用升变
        rules["back"] = rules_sheet[3][2].value == "c" # 启用悔棋
        rules["restrict_move_ne"] = rules_sheet[4][2].value == "c" # 限制连续3步以上移动中立棋子
        # 将打开的棋盘文件保存到设置中
        setting = load(SETTING_PATH)
        setting["lastly-load-map"] = filename
        write(SETTING_PATH,setting)
        return (chesses,map,rules)

def refresh_extension(extensions):
    setting = load(SETTING_PATH)
    setting["used-extensions"] = []
    for i in extensions:
        if i.use:
            setting["used-extensions"].append(i.name)
    write(SETTING_PATH,setting)
    static_data["used-extensions"] = setting["used-extensions"]

def refresh_color(value:str):
    setting = load(SETTING_PATH)
    setting["color-styles"] = value
    write(SETTING_PATH,setting)
    static_data["color-style-name"] = value
    static_data["colors"] = COLOR_STYLES[value]["colors"]

def set_record_path(path:str):
    setting = load(SETTING_PATH)
    setting["record-path"] = path
    write(SETTING_PATH,setting)
    static_data["record-path"] = path

def save_record(history:list[tuple[list[list],str]]):
    if not static_data["record-path"]:
        return
    print_chessboard = []
    for index,i in enumerate(history):
        print_chessboard.append([f"回合{index + 1}"])
        print_chessboard.extend([[k.name if k else " " for k in j] for j in i[0]])
    with open(static_data["record-path"],"w",newline = "") as wfile:
        writer = csv.writer(wfile)
        writer.writerows(print_chessboard)

def load_data():
    if os.path.exists(SETTING_PATH):
        setting = load(SETTING_PATH)
    else:
        setting = {
            "color-styles":"normal",
            "lastly-load-map":"",
            "record-path":"",
            "used-extensions":[]
        }
        write(SETTING_PATH,setting)
    static_data = {}
    try:
        static_data["color-styles"] = [(i,COLOR_STYLES[i]["name"]) for i in COLOR_STYLES]
        static_data["color-style-name"] = setting["color-styles"]
        static_data["colors"] = COLOR_STYLES[setting["color-styles"]]["colors"]
        static_data["shape-style"] = COMPONENT_STYLE
        static_data["lastly-load-map"] = setting["lastly-load-map"]
        static_data["record-path"] = setting["record-path"]
        static_data["used-extensions"] = setting["used-extensions"]
    except: # setting文件格式错误
        os.remove(SETTING_PATH)
        static_data = load_data()
    return static_data

static_data = load_data()
